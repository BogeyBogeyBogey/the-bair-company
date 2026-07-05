#!/usr/bin/env python3
"""
Build customer export bundles from HomePilot data.

Exports are deliberately dashboard-shaped: property overview, assessments,
interactions, recommendations, and a manifest. CSV files make the bundle easy
to inspect; XLSX gives customers the familiar Excel handoff.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from homepilot_snapshot import build_dashboard_snapshot
from homepilot_store import load_payload


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def is_dashboard_snapshot(data: dict[str, Any]) -> bool:
    return isinstance(data.get("tenant"), dict) and isinstance(data.get("properties"), list)


def load_dashboard_snapshot(
    path: Path,
    tenant_name: str | None = None,
    tenant_slug: str | None = None,
    modules: list[str] | None = None,
) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if is_dashboard_snapshot(data):
        return data
    payload = load_payload(path)
    return build_dashboard_snapshot(
        payload,
        tenant_name=tenant_name,
        tenant_slug=tenant_slug,
        enabled_modules=modules,
    )


def best_assessment(property_row: dict[str, Any]) -> tuple[str, dict[str, Any]]:
    assessments = property_row.get("assessments", {})
    if not assessments:
        return "", {}
    module_key = sorted(assessments, key=lambda key: assessments[key].get("score", 0), reverse=True)[0]
    return module_key, assessments[module_key]


def property_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for prop in snapshot.get("properties", []):
        best_module, best = best_assessment(prop)
        rows.append({
            "property_id": prop.get("id", ""),
            "address": prop.get("address", ""),
            "city": prop.get("city", ""),
            "lat": prop.get("lat", ""),
            "lon": prop.get("lon", ""),
            "status": prop.get("status", ""),
            "next_action": prop.get("nextAction", ""),
            "estimated_value": prop.get("estimatedValue", ""),
            "best_module": best_module,
            "best_score": best.get("score", ""),
            "best_grade": best.get("grade", ""),
            "tags": "; ".join(prop.get("tags", [])),
        })
    return rows


def assessment_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for prop in snapshot.get("properties", []):
        for module_key, assessment in prop.get("assessments", {}).items():
            rows.append({
                "property_id": prop.get("id", ""),
                "address": prop.get("address", ""),
                "module_key": module_key,
                "score": assessment.get("score", ""),
                "grade": assessment.get("grade", ""),
                "confidence": assessment.get("confidence", ""),
                "label": assessment.get("label", ""),
                "metrics_json": json.dumps(assessment.get("metrics", {}), ensure_ascii=False, sort_keys=True),
                "evidence_count": len(assessment.get("evidence", [])),
            })
    return rows


def interaction_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for prop in snapshot.get("properties", []):
        for interaction in prop.get("interactions", []):
            rows.append({
                "property_id": prop.get("id", ""),
                "address": prop.get("address", ""),
                "date": interaction.get("date", ""),
                "type": interaction.get("type", ""),
                "detail": interaction.get("detail", ""),
            })
    return rows


def recommendation_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"rank": index + 1, "recommendation": recommendation}
        for index, recommendation in enumerate(snapshot.get("recommendations", []))
    ]


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else ["empty"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    header_fill = PatternFill("solid", fgColor="E8EEF0")
    header_font = Font(bold=True)

    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name[:31])
        fieldnames = list(rows[0].keys()) if rows else ["empty"]
        sheet.append(fieldnames)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            sheet.append([row.get(field, "") for field in fieldnames])
        sheet.freeze_panes = "A2"
        for column_index, field in enumerate(fieldnames, start=1):
            max_len = max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows])
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 12), 48)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return True


def build_export_bundle(
    snapshot: dict[str, Any],
    output_dir: Path,
    include_xlsx: bool = True,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    sheets = {
        "properties": property_rows(snapshot),
        "assessments": assessment_rows(snapshot),
        "interactions": interaction_rows(snapshot),
        "recommendations": recommendation_rows(snapshot),
    }

    files: dict[str, str] = {}
    for name, rows in sheets.items():
        csv_path = output_dir / f"{name}.csv"
        write_csv(csv_path, rows)
        files[f"{name}_csv"] = str(csv_path)

    xlsx_written = False
    if include_xlsx:
        xlsx_path = output_dir / "homepilot_export.xlsx"
        xlsx_written = write_xlsx(xlsx_path, sheets)
        if xlsx_written:
            files["xlsx"] = str(xlsx_path)

    manifest = {
        "tenant": snapshot.get("tenant", {}),
        "summary": snapshot.get("summary", {}),
        "exported_at": utc_now(),
        "files": files,
        "xlsx_written": xlsx_written,
    }
    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    files["manifest"] = str(manifest_path)
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Build HomePilot customer export bundle")
    parser.add_argument("--json", required=True, type=Path, help="Canonical payload or dashboard snapshot JSON")
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument("--tenant-name", default="")
    parser.add_argument("--tenant-slug", default="")
    parser.add_argument("--module", dest="modules", action="append", default=None)
    parser.add_argument("--no-xlsx", action="store_true", help="Write CSV files only")
    args = parser.parse_args()

    snapshot = load_dashboard_snapshot(
        args.json,
        tenant_name=args.tenant_name or None,
        tenant_slug=args.tenant_slug or None,
        modules=args.modules,
    )
    manifest = build_export_bundle(snapshot, args.out_dir, include_xlsx=not args.no_xlsx)
    print(json.dumps(manifest, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
